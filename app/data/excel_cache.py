import os
import time
from threading import Lock
from openpyxl import load_workbook
from app.core.config import settings

class ExcelCache:
    _cache = {}
    _cache_lock = Lock()
    _write_locks = {}
    _ttl = settings.EXCEL_CACHE_TTL

    @classmethod
    def get_workbook(cls, file_path):
        now = time.time()
        with cls._cache_lock:
            cached = cls._cache.get(file_path)

            if cached:
                wb, last_load, last_mtime = cached
                current_mtime = os.path.getmtime(file_path)
                if now - last_load > cls._ttl or current_mtime != last_mtime:
                    wb = load_workbook(file_path)
                    cls._cache[file_path] = (wb, now, current_mtime)
                return cls._cache[file_path][0]

            wb = load_workbook(file_path)
            mtime = os.path.getmtime(file_path)
            cls._cache[file_path] = (wb, now, mtime)
            return wb

    @classmethod
    def invalidate(cls, file_path):
        with cls._cache_lock:
            if file_path in cls._cache:
                del cls._cache[file_path]

    @classmethod
    def acquire_write_lock(cls, file_path):
        with cls._cache_lock:
            if file_path not in cls._write_locks:
                cls._write_locks[file_path] = Lock()
            lock = cls._write_locks[file_path]
        lock.acquire()

    @classmethod
    def release_write_lock(cls, file_path):
        with cls._cache_lock:
            if file_path in cls._write_locks:
                cls._write_locks[file_path].release()
