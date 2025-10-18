import os
from openpyxl import load_workbook, Workbook
from datetime import datetime
from app.core.config import settings
from app.utils.logger import logger
from app.data.excel_cache import ExcelCache

class ExcelAdapter:
    VOTERS_SHEET = "Voters_Data"
    USERS_SHEET = "Users"
    UPDATES_SHEET = "Voter_Updates"
    BOOTH_SUMMARY_SHEET = "Booth_Summary"

    # ✅ Full expanded headers for Voters sheet
    VOTER_HEADERS = [
        "ConstituencyID", "ConstituencyName", "StateName", "BlockName",
        "PanchayatName", "BoothID", "BoothNumber", "BoothLocation",
        "VoterEPIC", "SerialNoInList",
        "Voter_fName", "Voter_lName", "Voter_fName_Hin", "Voter_lName_Hin",
        "Relation", "Guardian_fName", "Guardian_lName",
        "Guardian_fName_Hin", "Guardian_lName_Hin",
        "HouseNo", "Gender", "DOB", "Age", "Mobile", "EmailId",
        "LastVotedParty", "VotingPreference", "CertaintyOfVote", "VoteType", "Availability",
        "Religion", "Category", "OBCSubtype", "Caste", "LanguagePref",
        "EducationLevel", "EmploymentStatus", "GovtJobType", "GovtJobGroup", "JobRole",
        "MonthlySalaryRange", "PrivateJobRole", "PrivateSalaryRange",
        "SelfEmployedService", "BusinessType", "BusinessTurnoverRange",
        "GigWorkerRole", "ResidingIn", "Feedback", "PartNo", "OtherCity", "PermanentInBihar", "Migrated",
        "VerificationStatus", "CreatedAt", "UpdatedAt",
        # Additional fields
        "AdditionalComments", "AddressNotes", "AddressProof", "Area", "BusinessName",
        "BusinessTypeOther", "CasteOther", "CommunicationLanguage", "CommunityDetails",
        "CommunityParticipation", "CompanyName", "CropType", "CurrentLocation",
        "CustomRelationName", "DataConsent", "DevelopmentSuggestions",
        "DigitalCreatorCategory", "DigitalCreatorChannelName", "DigitalCreatorContentType",
        "DigitalCreatorFollowers", "DigitalCreatorIncome", "DigitalCreatorOtherCategory",
        "DigitalCreatorOtherPlatform", "DigitalCreatorPlatform", "FamilyContactNumber",
        "FamilyContactPerson", "FamilyHeadId", "FamilyRelation", "FamilyVotesTogether",
        "FirstTimeVoter", "GovtSchemes", "HouseType", "InfluencedByLeaders",
        "IsPartyWorker", "IssuesFaced", "LandHolding", "Landmark", "LanguageOther",
        "MigrationReason", "MlaSatisfaction", "MostImportantIssue", "OtherIssues",
        "PartyWorkerOtherParty", "PartyWorkerParty", "PinCode", "PostOffice",
        "SalaryRange", "Street", "UnemploymentReason", "VillageWard",
        "WorkExperience", "YearsSinceMigration"
    ]

    def __init__(self, constituency_file):
        self.file_path = os.path.join(settings.EXCEL_DIR, constituency_file)
        if not os.path.exists(self.file_path):
            self._create_new_file()
        self._load_workbook()

    def _create_new_file(self):
        logger.info(f"Creating new Excel file: {self.file_path}")
        wb = Workbook()

        # Create voters sheet with expanded headers
        ws_voters = wb.active
        ws_voters.title = self.VOTERS_SHEET
        ws_voters.append(self.VOTER_HEADERS)

        # Users sheet
        ws_users = wb.create_sheet(self.USERS_SHEET)
        ws_users.append([
            "UserID", "Username", "Role", "FullName", "Phone", "AssignedBoothIDs", "PasswordHash", "Email", "CreatedBy", "ParentID", "AssignedConstituencyIDs"
        ])

        # Voter updates sheet
        ws_updates = wb.create_sheet(self.UPDATES_SHEET)
        ws_updates.append(["UpdateID", "VoterEPIC", "UserID", "Changes", "CreatedAt"])

        # Booth summary sheet
        ws_booth_summary = wb.create_sheet(self.BOOTH_SUMMARY_SHEET)
        ws_booth_summary.append([
            "BoothID", "ConstituencyID", "TotalVoters", "MaleVoters", 
            "FemaleVoters", "OtherGenderVoters", "VotingPreferenceCounts",
            "ReligionCounts", "CategoryCounts", "EducationCounts", 
            "EmploymentCounts", "AgeGroupCounts", "LastUpdated"
        ])

        wb.save(self.file_path)

    def _load_workbook(self):
        self.wb = ExcelCache.get_workbook(self.file_path)
        self.ws_voters = self.wb[self.VOTERS_SHEET]
        self.ws_users = self.wb[self.USERS_SHEET]
        self.ws_updates = self.wb[self.UPDATES_SHEET]

    # -------- VOTERS --------
    def get_voters(self, booth_ids=None, constituency_id=None):
        rows = list(self.ws_voters.iter_rows(min_row=2, values_only=True))
        headers = [cell.value for cell in self.ws_voters[1]]
        data = [dict(zip(headers, row)) for row in rows]
        if booth_ids:
            data = [v for v in data if str(v.get("BoothID")) in booth_ids]
        if constituency_id:
            data = [v for v in data if v.get("ConstituencyID") == constituency_id]

        # print(data[0])
        return data

    def update_voter(self, epic_id, changes, user_id):
        # Acquire write lock for safe concurrent writes
        ExcelCache.acquire_write_lock(self.file_path)
        try:
            headers = [cell.value for cell in self.ws_voters[1]]
            epic_col = headers.index("VoterEPIC") + 1

            for row in range(2, self.ws_voters.max_row + 1):
                if self.ws_voters.cell(row=row, column=epic_col).value == epic_id:
                    old_data = {}
                    for key, value in changes.items():
                        if key in headers:
                            col = headers.index(key) + 1
                            old_value = self.ws_voters.cell(row=row, column=col).value
                            old_data[key] = old_value
                            self.ws_voters.cell(row=row, column=col, value=value)
                    audit_record = {
                        "old_values": old_data,
                        "new_values": changes
                    }
                    self._log_update(epic_id, user_id, audit_record)
                    self._save()
                    return True
            return False
        finally:
            ExcelCache.release_write_lock(self.file_path)

    # -------- USERS --------
    def get_users(self):
        rows = list(self.ws_users.iter_rows(min_row=2, values_only=True))
        headers = [cell.value for cell in self.ws_users[1]]
        return [dict(zip(headers, row)) for row in rows]

    # -------- UPDATES --------
    def _log_update(self, epic_id, user_id, changes):
        next_id = self.ws_updates.max_row
        self.ws_updates.append([
            next_id, epic_id, user_id, str(changes),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])

    # -------- SAVE --------
    def _save(self):
        self.wb.save(self.file_path)
        ExcelCache.invalidate(self.file_path)
